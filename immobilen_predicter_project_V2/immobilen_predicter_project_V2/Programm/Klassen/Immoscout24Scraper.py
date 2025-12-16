from playwright.sync_api import sync_playwright
import re
import os
import openpyxl
import pandas as pd
# import signal
# import sys


RED = '\033[31m'
BLUE = '\033[34m'
YELLOW = '\033[33m'
GREEN = '\033[32m'
RESET = '\033[0m'

class ImmoscoutScraper:
    def __init__(self, url, current_page_number=1):
        self.url = url
        self.current_page_number = current_page_number
        self.browser = None
        self.context = None
        self.page = None
        self.playwright = None
        self.verarbeitete_offers = 0
        self.captcha = True
        self.zustands_mapping = {
            'gepflegt': 'well_kept',
            'massivhaus, Gepflegt': 'well_kept',
            'renoviert / saniert': 'refurbished',
            'renoviert': 'refurbished',
            'saniert': 'refurbished',
            'erstbezug': 'first_time_use',
            'vollständig renoviert': 'fully_renovated',
            'neuwertig': 'mint_condition',
            'neubau, neuwertig': 'mint_condition',
            'Erstbezug nach Sanierung': 'first_time_use_after_refurbishment',
            'modernisiert': 'modernized',
            'verhandelbar': 'negotiable', 
            'renovierungsbedürftig': 'need_of_renovation',
            'Nach Vereinbarung': ''
        }

    # def signal_handler(self, sig, frame):
    #     print("\nAbbruch erkannt – versuche zu speichern...")
    #     if self.offers_data:
    #         try:
    #             self.speichern_daten_from_list(self.offers_data, dateiname='immoscout_daten.csv', format='csv', path='.')
    #             print("Daten erfolgreich gespeichert.")
    #         except Exception as e:
    #             print(f"Fehler beim Speichern im Abbruch-Handler: {e}")
    #     else:
    #         print("Keine Daten zum Speichern.")
    #     sys.exit(0)

    # def get_current_url(self):
    #     print(self.page.url)
    #     return self.page.url


    def start_browser(self):
        self.playwright = sync_playwright().start()
        self.browser = self.playwright.chromium.launch(headless=False)
        self.context = self.browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36",
            permissions=["geolocation"],
            locale="de-DE"
        )
        self.page = self.context.new_page()
        self.add_stealth(self.page)

    def stop_browser(self):
        if self.browser:
            self.browser.close()
        if self.playwright:
            self.playwright.stop()

    def collect_expose_ids(self):
        self.page.goto(self.url)
        #time.sleep(random.uniform(2, 5))

        if(self.captcha):
            print("Bitte löse das Captcha und drücke Enter, um fortzufahren...")
            input()
            self.captcha = False

        expose_ids = set()
        try:
            self.page.wait_for_selector('a[data-exp-id]', timeout=15000)
            links = self.page.locator('a[data-exp-id]').all()

            for link in links:
                exp_id = link.get_attribute('data-exp-id')
                if exp_id:
                    expose_ids.add(exp_id)

            return list(expose_ids)
        except:
            self.captcha = True
            return self.collect_expose_ids()


    def add_stealth(self, page):
        # Mausbewegungen simulieren
        page.mouse.move(200, 200)
        page.mouse.down()
        page.mouse.move(300, 300, steps=5)
        page.mouse.up()

        # Weitere Stealth-Techniken
        page.evaluate("navigator.webdriver = false;")
        page.evaluate("""
            Object.defineProperty(navigator, 'headless', {get: () => false});
        """)

    def get_next_page_url(self):
        if "?pagenumber=" in self.url:
            base_url = self.url.split("?pagenumber=")[0]
        else:
            base_url = self.url

        new_url = f"{base_url}?pagenumber={self.current_page_number}"
        #print(f"Neue URL: {new_url}")
        return new_url

    def get_inhalte_from_offer(self, expose_id):
        offer_url = f"https://www.immobilienscout24.de/expose/{expose_id}"
        #print(f"Öffne Angebot: {offer_url}")
        self.page.goto(offer_url, timeout=60000)
        #time.sleep(random.uniform(1, 3))
        p = self.page
        wohnflaeche = self.extrahiere_wohnflaeche(p)
        zimmeranzahl = self.extrahiere_zimmeranzahl(p)
        baujahr = self.extrahiere_baujahr(p)
        zustand = self.extrahiere_zustand(p)
        kueche = self.extrahiere_kueche(p)
        terrasse_balkon = self.extrahiere_balkon_terrasse(p)
        aufzug = self.extrahiere_aufzug(p)
        garten = self.extrahiere_garten(p)
        keller = self.extrahiere_keller(p)
        etage = self.extrahiere_etage(p)
        kaltmiete = self.extrahiere_kaltmiete(p)
        warmmiete = self.extrahiere_warmmiete(p)
        energie_effizienz = self.extrahiere_energy_efficiency_class(p)
        bundesland = self.extrahiere_bundesland(p)
        stadt, plz = self.extrahiere_stadt_plz(p)
        stadtteil = self.extrahiere_stadtteil(p)




        return wohnflaeche, zimmeranzahl, baujahr, zustand, kueche, terrasse_balkon, aufzug, garten, keller, etage, kaltmiete, warmmiete, energie_effizienz, bundesland, stadt, stadtteil, plz

    def extrahiere_wohnflaeche(self, page):
        try:
            page.wait_for_selector("dd.is24qa-wohnflaeche-ca", timeout=100)
            element = page.locator("dd.is24qa-wohnflaeche-ca").first
            text = element.text_content().strip()
            match = re.search(r'[\d.,]+', text)
            if match:
                #print(float(match.group(0).replace(',', '.')))
                return float(match.group(0).replace(',', '.'))
        except Exception as e:
            #print(f"Fehler bei der Extraktion der Wohnfläche: {e}")
            return None
    
    def extrahiere_zimmeranzahl(self, page):
        try:
            page.wait_for_selector("dd.is24qa-zimmer", timeout=100)
            element = page.locator("dd.is24qa-zimmer").first
            text = element.text_content().strip()
            match = re.search(r'[\d.,]+', text)
            if match:
                return float(match.group(0).replace(',', '.'))
        except Exception as e:
            #print(f"Fehler bei der Extraktion der Zimmer: {e}")
            return None
    
    def extrahiere_baujahr(self, page):
        try:
            page.wait_for_selector("dd.is24qa-baujahr", timeout=100)
            element = page.locator("dd.is24qa-baujahr").first
            text = element.text_content().strip()
            return text
        except Exception as e:
            #print(f"Fehler bei der Extraktion vom Baujahr: {e}")
            return None
    
    def extrahiere_zustand(self, page):
        try:
            page.wait_for_selector("dd.is24qa-objektzustand", timeout=100)
            element = page.locator("dd.is24qa-objektzustand").first
            text = element.text_content().strip()
            text_englisch = self.zustands_mapping.get(text.lower(), None)
            return text_englisch
        except Exception as e:
            #print(f"Fehler bei der Extraktion vom Baujahr: {e}")
            return None
    
    def extrahiere_kueche(self, page):
        try:
            page.wait_for_selector("span.is24qa-einbaukueche-label", timeout=100)
            element = page.locator("span.is24qa-einbaukueche-label").first
            text = element.text_content().strip().lower()
            if "küche" in text:
                return True
        except Exception as e:
            #print(f"Fehler bei der Extraktion der Küche: {e}")
            return False
    
    def extrahiere_balkon_terrasse(self, page):
        try:
            page.wait_for_selector("span.is24qa-balkon-terrasse-label", timeout=100)
            element = page.locator("span.is24qa-balkon-terrasse-label").first
            text = element.text_content().strip().lower()
            if "balkon" and "terrasse" in text:
                return True
        except Exception as e:
            #print(f"Fehler bei der Extraktion des Balkon: {e}")
            return False
    
    def extrahiere_aufzug(self, page):
        try:
            page.wait_for_selector("span.is24qa-personenaufzug-label", timeout=100)
            element = page.locator("span.is24qa-personenaufzug-label").first
            text = element.text_content().strip().lower()
            if "aufzug" in text:
                return True
        except Exception as e:
            #print(f"Fehler bei der Extraktion des Afzugs: {e}")
            return False
        
    def extrahiere_garten(self, page):
        try:
            page.wait_for_selector("span.is24qa-garten-label", timeout=100)
            element = page.locator("span.is24qa-garten-label").first
            text = element.text_content().strip().lower()
            if "garten" in text:
                return True
        except Exception as e:
            #print(f"Fehler bei der Extraktion des Afzugs: {e}")
            return False

    def extrahiere_keller(self, page):
        try:
            page.wait_for_selector("span.is24qa-keller-label", timeout=100)
            element = page.locator("span.is24qa-keller-label").first
            text = element.text_content().strip().lower()
            if "keller" in text:
                return True
        except Exception as e:
            #print(f"Fehler bei der Extraktion des Afzugs: {e}")
            return False
    
    def extrahiere_etage(self, page):
        try:
            page.wait_for_selector("dd.is24qa-etage", timeout=100)
            element = page.locator("dd.is24qa-etage").first
            text = element.text_content().strip().lower()
            if 'von' in text:
                etage = text.split("von")[0].strip()
                return etage
            else:
                return None
            
        except Exception as e:
            #print(f"Fehler bei der Extraktion der Etage: {e}")
            return None
    
    def extrahiere_kaltmiete(self, page):
        try:
            page.wait_for_selector("dd.is24qa-kaltmiete", timeout=100)
            element = page.locator("dd.is24qa-kaltmiete").first
            zahl = element.text_content().replace(".", "").strip()
            zahl = re.sub(r"€.*", "", zahl).strip()
            zahl = zahl.replace(",", ".")
            return zahl
            
        except Exception as e:
           # print(f"Fehler bei der Extraktion der Kaltmiete: {e}")
            return None
    
    def extrahiere_warmmiete(self, page):
        try:
            page.wait_for_selector("dd.is24qa-gesamtmiete", timeout=100)
            element = page.locator("dd.is24qa-gesamtmiete").first
            zahl = element.text_content().replace(".", "").strip()
            zahl = re.sub(r"€.*", "", zahl).strip()
            zahl = zahl.replace(",", ".")
            return zahl
            
        except Exception as e:
            #print(f"Fehler bei der Extraktion der Warmmiete: {e}")
            return None
        
    def extrahiere_energy_efficiency_class(self, page):
        try:
            page.wait_for_selector("span.energy-efficiency-class img[alt]", timeout=100)
            element = page.locator("span.energy-efficiency-class img[alt]").first
            alt_text = element.get_attribute('alt')
            #print(alt_text)
            if alt_text == "A_PLUS": return "A+"
            elif alt_text in ["A", "B", "C", "D", "E", "F", "G"]: return alt_text
            else: return None
            
        except Exception as e:
            #print(f"Fehler bei der Extraktion der Energieeffizienz: {e}")
            return None
        
    def extrahiere_stadt_plz(self, page):
        try:
            page.wait_for_selector("span.zip-region-and-country", timeout=100)
            element = page.locator("span.zip-region-and-country").first
            text = element.text_content().strip()

            match = re.search(r'(\d{5})\s+(.+)', text)
            if match:
                plz = match.group(1)
                stadt = match.group(2)
                return stadt, plz
            return None, None
        except Exception as e:
            #print(f"Fehler bei der Extraktion der Stadt und PLZ: {e}")
            return None, None
        
    def extrahiere_stadtteil(self, page):
        try:
            page.wait_for_selector("span.zip-region-and-country", timeout=100)
            element = page.locator("span.zip-region-and-country").first
            stadtteil = element.text_content().split(',')[0].strip()

            stadtteil = re.sub(r'\d+', '', stadtteil).strip()
            #print(stadtteil)
            return stadtteil
        except Exception as e:
            # print(f"Fehler bei der Extraktion der Stadt und PLZ: {e}")
            return None
        
    def extrahiere_bundesland(self, page):
        try:
            page.wait_for_selector("li.breadcrumbs-item a", timeout=100)
            link = page.locator("li.breadcrumbs-item a").nth(1)
            href = link.get_attribute("href")
            if href:
                bundesland = href.split("/")[3]
                #print(bundesland)
                return bundesland

        except Exception as e:
            #print(f"Fehler bei der Extraktion der Stadt und PLZ: {e}")
            return None
        
    def speichern_daten_from_list(self, offers_data, dateiname, format='csv', path='.'):
        full_path = os.path.join(path, dateiname)
        os.makedirs(path, exist_ok=True)

        df = pd.DataFrame(offers_data)

        if format == 'csv':
            if os.path.exists(full_path):
                df.to_csv(full_path, mode='a', header=False, index=False, encoding='utf-8')
                print(f"{GREEN}Datei im Pfad:{RESET} {full_path} {GREEN}wurde erweitert{RESET}")
            else:
                df.to_csv(full_path, index=False, encoding='utf-8')
                print(f"{GREEN}Datei als .csv gspeichert{RESET}")

        elif format == 'xlsx':
            if os.path.exists(full_path):
                wb = openpyxl.load_workbook(full_path)
                sheet = wb.active
                start_row = sheet.max_row + 1

                for r_idx, row in df.iterrows():
                    for c_idx, value in enumerate(row):
                        sheet.cell(row=start_row + r_idx, column=c_idx + 1, value=value)
                wb.save(full_path)
                print(f"{GREEN}Datei in: {RESET}{full_path} {GREEN}wurde erweitert{RESET}")
            else:
                df.to_excel(full_path, index=False, engine='openpyxl')
                print(f"Datei als .xlsx {GREEN}gspeichert{RESET}")


    def lade_bereits_gespeicherte_offer_ids(self, path, dateiname):
        full_path = os.path.join(path, dateiname)

        if not os.path.exists(full_path):
            return set()
        
        try:
            if full_path.endswith('.csv'):
                df = pd.read_csv(full_path)
            else:
                df = pd.read_excel(full_path)

            if 'Offer_id' in df.columns:
                return set(df['Offer_id'].astype(str))
            else:
                return set()
        except Exception as e:
            print(f"Fehler beim Einlesen bestehender Datei: {e}")
            return set()
        
    def scrape_multiple_offers(self, add_number_of_offers=10, dateiname='immoscout_daten.csv', format='csv', path='.'):
        
        # signal.signal(signal.SIGINT, self.signal_handler)
        # signal.signal(signal.SIGINT, self.signal_handler)


        offers_data = []
        neu_scraped_ids = set()
        already_scraped_ids = self.lade_bereits_gespeicherte_offer_ids(path, dateiname)
        offers_scraped = 0

        try:
            while offers_scraped < add_number_of_offers:
                try:
                    print(f"Verarbeite Seite {YELLOW}{self.current_page_number}{RESET}...")
                        

                    ids = self.collect_expose_ids()
                    if not ids:
                        print("Keine weiteren Angebote gefunden.")
                        break

                    for offer_id in ids:
                        if offers_scraped >= add_number_of_offers:
                            break

                        if offer_id in already_scraped_ids or offer_id in neu_scraped_ids:
                            print(f"Angebot {offer_id} bereits vorhanden. Überspringe...")
                            continue

                        print(f"Verarbeite {BLUE}neues{RESET} Angebot ID: {offer_id}")

                        try:
                            data = self.get_inhalte_from_offer(offer_id)
                            if data:
                                wohnflaeche, zimmer, baujahr, zustand, kueche, terrasse_balkon, aufzug, garten, keller, etage, kaltmiete, warmmiete, energieeffizienz, bundesland, stadt, stadtteil, plz = data
                                offers_data.append({
                                    'Wohnfläche': wohnflaeche,
                                    'Zimmeranzahl': zimmer,
                                    'Baujahr': baujahr,
                                    'Zustand': zustand,
                                    'Küche': kueche,
                                    'Balkon / Terrasse': terrasse_balkon,
                                    'Aufzug': aufzug,
                                    'Garten': garten,
                                    'Keller': keller,
                                    'Etage': etage,
                                    'Kaltmiete': kaltmiete,
                                    'Warmmiete': warmmiete,
                                    'Energieeffizienzklasse': energieeffizienz,
                                    'Bundesland': bundesland,
                                    'Stadt': stadt,
                                    'Stadtteil': stadtteil,
                                    'PLZ': plz,
                                    'Offer_id': offer_id
                                })
                                neu_scraped_ids.add(offer_id)
                                offers_scraped += 1
                                self.verarbeitete_offers += 1
                                print(f"{YELLOW}{offers_scraped}{RESET}/{add_number_of_offers} Verarbeitet")
                        except Exception as e:
                            print(f"Fehler beim Verarbeiten von Angebot {offer_id}: {e}")

                    self.current_page_number += 1
                    self.url = self.get_next_page_url()

                except Exception as e:
                    print(f"Fehler auf Seite {self.current_page_number}: {e}")
                    self.current_page_number += 1 
                    continue

        except Exception as e:
            print(f"Kritischer Fehler: {e}")

        finally:
            if offers_data:
                print(f"{GREEN}Speichere{RESET} gesammelte Daten...")
                try:
                    self.speichern_daten_from_list(offers_data, dateiname=dateiname, format=format, path=path)
                except Exception as e:
                    print(f"Fehler beim Speichern: {e}")
            else:
                print("Keine neuen Daten zum Speichern.")



if __name__ == "__main__":

    pagenumber = 1
    url = f"https://www.immobilienscout24.de/Suche/de/wohnung-mieten?pagenumber={pagenumber}"
    scraper = ImmoscoutScraper(url, pagenumber)
    scraper.start_browser()

    try:
        if __name__ == "__main__":
            scraper.scrape_multiple_offers(add_number_of_offers=110000, dateiname='immoscout_daten_2025_12.xlsx', format='xlsx', path='../Tabellen')

    finally:
        
        scraper.stop_browser()
