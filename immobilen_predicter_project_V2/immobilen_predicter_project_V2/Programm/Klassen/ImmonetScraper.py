import requests 
from bs4 import BeautifulSoup  
import re
import pandas as pd
import os
import openpyxl
import random
import time

RED = '\033[31m'
BLUE = '\033[34m'
YELLOW = '\033[33m'
GREEN = '\033[32m'
RESET = '\033[0m'

class ImmonetScraper:
    def __init__(self, delay=2.0, url='https://www.immonet.de/classified-search?distributionTypes=Rent&estateTypes=Apartment&locations=AD02DE1&page=1'):
        self.url = url
        self.delay = delay
        self.reconnect_interval = random.randint(11, 17)

        self.user_agents = [
            # Chrome (Windows)
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36',
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36',
            'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36',
            
            # Chrome (Linux)
            'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36',
            
            # Firefox (Windows)
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:124.0) Gecko/20100101 Firefox/124.0',
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:123.0) Gecko/20100101 Firefox/123.0',
            'Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:122.0) Gecko/20100101 Firefox/122.0',

            # Firefox (Linux)
            'Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:124.0) Gecko/20100101 Firefox/124.0',
            
            # Safari (Mac)
            'Mozilla/5.0 (Macintosh; Intel Mac OS X 13_3) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.0 Safari/605.1.15',
            'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/16.0 Safari/605.1.15',
            
            # Edge (Windows)
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36 Edg/124.0.0.0',
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36 Edg/123.0.0.0',

            # Mobile (Android)
            'Mozilla/5.0 (Linux; Android 13; Pixel 7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Mobile Safari/537.36',
            'Mozilla/5.0 (Linux; Android 11; SM-A715F) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Mobile Safari/537.36',
            
            # Mobile (iPhone)
            'Mozilla/5.0 (iPhone; CPU iPhone OS 17_0 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.0 Mobile/15E148 Safari/604.1',
            'Mozilla/5.0 (iPhone; CPU iPhone OS 16_5 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/16.5 Mobile/15E148 Safari/604.1'
        ]

        self.header = self.create_header()
        self.heizung_mapping = {
            'Fußbodenheizung': 'floor_heating',
            'Zentralheizung': 'central_heating',
            'Fernwärme': 'district_heating',
            'Gasheizung': 'gas_heating',
            'Ölheizung': 'oil_heating',
            'Kombinierte Heizungs- und Stromversorgung': 'combined_heat_and_power_plant',
            'Elektroheizung': 'electric_heating',
            'Etagenheizung': 'self_contained_central_heating',
            'Ofen': 'stove_heating',
            'Pelletheizung': 'wood_pellet_heating',
            'Wärmepumpe': 'heat_pump',
            'Etagenheizung: Fußbodenheizung': 'floor_heating'

        }
        self.zustands_mapping = {
            'gepflegt': 'well_kept',
            'massivhaus, Gepflegt': 'well_kept',
            'renoviert / saniert': 'refurbished',
            'renoviert': 'refurbished',
            'erstbezug': 'first_time_use',
            'vollständig renoviert': 'fully_renovated',
            'neuwertig': 'mint_condition',
            'neubau, neuwertig': 'mint_condition',
            'Erstbezug nach Sanierung': 'first_time_use_after_refurbishment',
            'modernisiert': 'modernized',
            'verhandelbar': 'negotiable', 
            'renovierungsbedürftig': 'need_of_renovation'
        }
        self.session = self.create_session()
        self.verarbeitete_offers = 0

    def create_header(self):
        return {
            'User-Agent': random.choice(self.user_agents),
            'Accept-Language': 'de-DE,de;q=0.9,en-US;q=0.8,en;q=0.7',
            'Accept-Encoding': 'gzip, deflate, br',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
            'Connection': 'keep-alive',
            'Referer': 'https://www.immonet.de/'
        }
    
    def create_session(self):
        session = requests.Session()
        session.headers.update(self.create_header())
        return session

    def get_links_from_site(self, url):
        response = self.session.get(url)

        if response.status_code != 200:
            print(f"Fehler beim Laden der Seite: {url} (Status {response.status_code})")
            return []

        soup = BeautifulSoup(response.content, 'html.parser')
        expose_elements = soup.find_all('div', class_='css-79elbk')
        expose_ids = []  

        for element in expose_elements:
            data_testid = element.get('data-testid')
            if data_testid and data_testid.startswith('classified-card-mfe-'):
                expose_id = data_testid.replace('classified-card-mfe-', '') 
                expose_ids.append(expose_id)

        return expose_ids
    
    def get_inhalte_from_offer(self, id):
       
        offer_url = f"https://www.immonet.de/expose/{id}" 
        response = requests.get(offer_url, headers=self.header)

        if response.status_code == 200:
            self.soup = BeautifulSoup(response.content, 'html.parser')

            baujahr = scraper.extrahiere_baujahr()
            wohnflaeche = scraper.extrahiere_wohnflaeche()
            zimmer = scraper.extrahiere_zimmeranzahl()
            zustand =scraper.extrahiere_zustand()
            kueche = scraper.extrahiere_kueche()
            balkon = scraper.extrahiere_balkon()
            terrasse = scraper.extrahiere_terrasse()
            aufzug = scraper.extrahiere_aufzug()
            garten = scraper.extrahiere_garten()
            keller = scraper.extrahiere_keller()
            etage = scraper.extrahiere_etage()
            warmmiete, kaltmiete = scraper.extrahiere_miete()
            energieeffizienz = scraper.extrahiere_energieeffizienzklasse()
            bundesland = scraper.extrahiere_bundesland()
            stadt = scraper.extrahiere_stadt()
            stadtteil = scraper.extrahiere_stadtteil()
            plz = scraper.extrahiere_plz()


            return wohnflaeche, zimmer, baujahr, zustand, kueche, balkon, terrasse, aufzug, garten, keller, etage, kaltmiete, warmmiete, energieeffizienz, bundesland, stadt, stadtteil, plz
        else:
            print(f"Fehler beim Abrufen der Detailseite für Angebot {id}")
            return False
        
    def extrahiere_wohnflaeche(self):
        if not self.soup:
            return None

        spans = self.soup.find_all("span", class_="css-2bd70b")

        for span in spans:
            text = span.get_text(strip=True).replace(",", ".")
            if "m²" in text:
                try:
                    return float(text.replace("m²", "").strip())
                except ValueError:
                    return None
        return None
        
    def extrahiere_zimmeranzahl(self):
        if not self.soup:
            return None

        spans = self.soup.find_all("span", class_="css-2bd70b")

        for i, span in enumerate(spans):
            text = span.get_text(strip=True)
            if "m²" in text:
                if i > 0:
                    vorheriger_text = spans[i - 1].get_text(strip=True).replace(",", ".")
                    try:
                        return float(vorheriger_text)
                    except ValueError:
                        return None
        return None
    
    def extrahiere_baujahr(self):
        if not hasattr(self, 'soup') or not self.soup:
            return None
        baujahr_element = self.soup.find("span", {"data-testid": "aviv.CDP.Sections.Energy.Features.yearOfConstruction"})
        if baujahr_element:
            return baujahr_element.text.strip()
        return None

    def extrahiere_zustand(self):
        if not self.soup:
            return None
        zustands_element = self.soup.find("span", {"data-testid": "aviv.CDP.Sections.Energy.Features.state"})
        if zustands_element:
            zustands_element = zustands_element.text.strip()

            zustands_element_englisch = self.zustands_mapping.get(zustands_element.lower(), None)
            if zustands_element_englisch:
                return zustands_element_englisch
            
        return None
    
    def extrahiere_kueche(self):
        if not self.soup:
            return None
        kueche_element = self.soup.find("span", class_="css-1az3ztj", string=re.compile("küche", re.IGNORECASE))
        if kueche_element:
            return True
        return False
    
    def extrahiere_balkon(self):
        if not self.soup:
            return None
        balkon_element = self.soup.find("span", class_="css-1az3ztj", string=re.compile("Balkon", re.IGNORECASE))
        if balkon_element:
            return True
        return False
    
    def extrahiere_terrasse(self):
        if not self.soup:
            return None
        terassen_element = self.soup.find("span", class_="css-1az3ztj", string=re.compile("Terrasse", re.IGNORECASE))
        if terassen_element:
            return True
        return False
    
    def extrahiere_garten(self):
        if not self.soup:
            return None
        garten_element = self.soup.find("span", class_="css-1az3ztj", string=re.compile("Garten", re.IGNORECASE))
        if garten_element:
            return True
        return False
    
    def extrahiere_keller(self):
        if not self.soup:
            return None
        keller_element = self.soup.find("span", class_="css-1az3ztj", string=re.compile("Keller", re.IGNORECASE))
        if keller_element:
            return True
        return False
    
    def extrahiere_aufzug(self):
        if not self.soup:
            return None
        aufzug_element = self.soup.find("span", class_="css-1az3ztj", string=re.compile("Aufzug", re.IGNORECASE))
        if aufzug_element:
            return True
        return False
    
    def extrahiere_garten(self):
        if not self.soup:
            return None
        garten_element = self.soup.find("span", class_="css-1az3ztj", string=re.compile("Garten", re.IGNORECASE))
        if garten_element:
            return True
        return False
    
    def extrahiere_keller(self):
        if not self.soup:
            return None
        keller_element = self.soup.find("span", class_="css-1az3ztj", string=re.compile("Keller", re.IGNORECASE))
        if keller_element:
            return True
        return False
    
    def extrahiere_etage(self):
        if not self.soup:
            return None

        etage_element = self.soup.find("span", class_="css-1az3ztj", string=re.compile(r"(Erdgeschoss|\d+)\. Geschoss", re.IGNORECASE))
        if etage_element:
            text = etage_element.text.strip()
            match = re.search(r"(\d+)", text)
            if match:
                return int(match.group(1))
            
        etage_element = self.soup.find("span", class_="css-1az3ztj", string=re.compile("Erdgeschoss", re.IGNORECASE))
        if etage_element:
            return 0
        return None
        
    def extrahiere_miete(self):
        if not self.soup:
            return None, None
        
        warmmiete, kaltmiete = None, None

        kaltmiete_elements = self.soup.find_all("div", class_="css-y29352")
        for element in kaltmiete_elements:
            preis_text = element.find("span", class_="css-1gs73yw")
            if preis_text:
                preis_text = preis_text.get_text(strip=True)
                if "Kaltmiete" in element.get_text():
                    preis_zahl = preis_text.replace("€", "").replace(".", "").replace("\xa0", "").strip()
                    preis_zahl = re.sub(r"[^\d,\.]", "", preis_zahl)
                    preis_zahl = preis_zahl.replace(",", ".")
                    kaltmiete = float(preis_zahl)
                    break

        warmmiete_elements = self.soup.find_all("div", class_="css-cxt05v")
        for element in warmmiete_elements:
            preis_text = element.find("span", class_="css-2bd70b")
            if preis_text:
                preis_text = preis_text.get_text(strip=True)
                if "Warmmiete" in element.get_text():
                    preis_zahl = preis_text.replace("€", "").replace(".", "").replace("\xa0", "").strip()
                    preis_zahl = re.sub(r"[^\d,\.]", "", preis_zahl)
                    preis_zahl = preis_zahl.replace(",", ".")
                    warmmiete = float(preis_zahl)
                    break  

        return warmmiete, kaltmiete

    def extrahiere_energieeffizienzklasse(self):
        if not self.soup:
            return None 
        energieeffizienz_element = self.soup.find("div", {"data-testid": "aviv.CDP.Sections.Energy.Preview.EfficiencyClass"})
        if energieeffizienz_element:
            energieeffizienzklasse = energieeffizienz_element.get_text(strip=True)
            return energieeffizienzklasse
        else:
            return None
        
    def extrahiere_bundesland(self):
        if not self.soup:
            return None

        breadcrumb = self.soup.find("ol", class_="css-1ggu0ou")
        if breadcrumb:
            links = breadcrumb.find_all("a", class_="css-1q0e9w6")
            if len(links) > 2:
                text = links[2].get_text(strip=True)
                bundesland = text.replace("Wohnung zur Miete in ", "").strip()
                return bundesland
        return None
    
    def extrahiere_plz(self):
        if not self.soup:
            return None
        plz_element = self.soup.find("span", style="white-space:nowrap")
        if plz_element:
            text = plz_element.text.strip()
            match = re.search(r"\((\d{5})\)", text)
            if match:
                return match.group(1)
        return None

    def extrahiere_stadt(self):
        if not self.soup:
            return None
        stadt_element = self.soup.find("span", style="white-space:nowrap")
        if stadt_element:
            text = stadt_element.text.strip()
            match = re.search(r"^(.*)\s\((\d{5})\)$", text)
            if match:
                stadt_name = match.group(1).strip() 
                return stadt_name
        return None
    
    def extrahiere_stadtteil(self):
        if not self.soup:
            return None
        el = self.soup.find("div", class_="css-1ytyjyb")
        if el:
            text = el.get_text(separator=" ").strip()
            teile = text.split(",")
            if len(teile) == 3:
                return teile[1].strip()
            elif len(teile) == 2:
                return teile[0].strip()
        return None
    
    def lade_bereits_gespeicherte_offer_ids(self, path, dateiname):
        full_path = os.path.join(path, dateiname)

        if not os.path.exists(full_path):
            return set()
        
        try:
            if full_path.endswith('.csv'):
                df = pd.read_csv(full_path)
            else:
                df = pd.read_excel(full_path)

            if 'Offer_id' in df.columns:
                return set(df['Offer_id'].astype(str))
            else:
                return set()
        except Exception as e:
            print(f"Fehler beim Einlesen bestehender Datei: {e}")
            return set()
        
    def speichern_daten_from_list(self, offers_data, dateiname, format='csv', path='.'):
        full_path = os.path.join(path, dateiname)
        os.makedirs(path, exist_ok=True)

        df = pd.DataFrame(offers_data)

        if format == 'csv':
            if os.path.exists(full_path):
                df.to_csv(full_path, mode='a', header=False, index=False, encoding='utf-8')
                print(f"{GREEN}Datei im Pfad:{RESET} {full_path} {GREEN}wurde erweitert{RESET}")
            else:
                df.to_csv(full_path, index=False, encoding='utf-8')
                print(f"{GREEN}Datei als .csv gspeichert{RESET}")

        elif format == 'xlsx':
            if os.path.exists(full_path):
                wb = openpyxl.load_workbook(full_path)
                sheet = wb.active
                start_row = sheet.max_row + 1

                for r_idx, row in df.iterrows():
                    for c_idx, value in enumerate(row):
                        sheet.cell(row=start_row + r_idx, column=c_idx + 1, value=value)
                wb.save(full_path)
                print(f"{GREEN}Datei in: {RESET}{full_path} {GREEN}wurde erweitert{RESET}")
            else:
                df.to_excel(full_path, index=False, engine='openpyxl')
                print(f"{GREEN}Datei als .xlsx gspeichert{RESET}")

    def get_next_page_url(self, current_page_number):

        if f"&page=" in self.url:
            base_url = self.url.split("&page=")[0]
        else:
            base_url = self.url

        new_url = f"{base_url}&page={current_page_number}"
        print(f"Neue URL: {new_url}")
        
        return new_url
    
    def scrape_multiple_offers(self, add_number_of_offers=10, dateiname='immonet_daten.csv', format='csv', path='.', current_page_number=1):

            offers_data = []
            neu_scraped_ids = set()
            already_scraped_ids = self.lade_bereits_gespeicherte_offer_ids(path, dateiname)

            offers_scraped = 0

            try:
                while offers_scraped < add_number_of_offers:
                    try:
                        print(f"Verarbeite Seite {YELLOW}{current_page_number}{RESET}...")
                        if current_page_number == 1:
                            url = self.url
                        else:
                            url = self.get_next_page_url(current_page_number)

                        ids = self.get_links_from_site(url)
                        if not ids:
                            print("Keine weiteren Angebote gefunden.")
                            break

                        for offer_id in ids:
                            if offers_scraped >= add_number_of_offers:
                                break

                            if offer_id in already_scraped_ids or offer_id in neu_scraped_ids:
                                print(f"Angebot {offer_id} bereits vorhanden. Überspringe...")
                                continue
                            
                            print(f"Verarbeite {BLUE}neues{RESET} Angebot ID: {offer_id}")

                            try:
                                data = self.get_inhalte_from_offer(offer_id)
                                if data:
                                    wohnflaeche, zimmer, baujahr, zustand, kueche, balkon, terrasse, aufzug, garten, keller, etage, kaltmiete, warmmiete, energieeffizienz, bundesland, stadt, stadtteil, plz = data
                                    offers_data.append({
                                        'Wohnfläche': wohnflaeche,
                                        'Zimmeranzahl': zimmer,
                                        'Baujahr': baujahr,
                                        'Zustand': zustand,
                                        'Küche': kueche,
                                        'Balkon': balkon,
                                        'Terrasse': terrasse,
                                        'Aufzug': aufzug,
                                        'Garten': garten,
                                        'Keller': keller,
                                        'Etage': etage,
                                        'Kaltmiete': kaltmiete,
                                        'Warmmiete': warmmiete,
                                        'Energieeffizienzklasse': energieeffizienz,
                                        'Bundesland': bundesland,
                                        'Stadt': stadt,
                                        'Stadtteil': stadtteil,
                                        'PLZ': plz,
                                        'Offer_id': offer_id
                                    })
                                    offers_scraped += 1
                                    self.verarbeitete_offers += 1
                                    print(f"{YELLOW}{offers_scraped}{RESET}/{add_number_of_offers} Verarbeitet")
                            except Exception as e:
                                print(f"Fehler beim Verarbeiten von Angebot {offer_id}: {e}")


                        current_page_number += 1
                        
                    except Exception as e:
                        print(f"Fehler auf Seite {self.current_page_number}: {e}")
                        self.current_page_number += 1 
                        continue
            
            except Exception as e:
                print(f"Kritischer Fehler: {e}")

            finally:
                if offers_data:
                    print(f"{GREEN}Speichere{RESET} gesammelte Daten...")
                    try:
                        self.speichern_daten_from_list(offers_data, dateiname=dateiname, format=format, path=path)
                    except Exception as e:
                        print(f"Fehler beim Speichern: {e}")
                else:
                    print("Keine neuen Daten zum Speichern.")


if __name__ == "__main__":
    scraper = ImmonetScraper()
    format = 'xlsx'
    while True:
        scraper.scrape_multiple_offers(add_number_of_offers=15000, dateiname=f'immonet_daten.{format}', format=format, path='C:/temp/____Noah_Ordner/1py_Programme/Programme/Machine_Learning/Code/immobilen_predicter_project_V2/Programm/Tabellen')
        
        time.sleep(10)

